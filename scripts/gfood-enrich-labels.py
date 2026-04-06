#!/usr/bin/env python3
"""
Enrich G-Food article Excel: extract Contenance (normalized units), Libellé_nettoyé,
Marque, split multilingual segments, optional DeepL translations: origin → French, then
French → English / German / Italian.

Examples:
  python3 scripts/gfood-enrich-labels.py --input docs/file.xlsx --dry-run --limit 50
  python3 scripts/gfood-enrich-labels.py --input docs/file.xlsx --dry-run --dry-run-translate --limit 10
  python3 scripts/gfood-enrich-labels.py --input docs/file.xlsx --output docs/file-enriched.xlsx
  python3 scripts/gfood-enrich-labels.py -i docs/file.xlsx -o out.xlsx --stream-output --progress-every 200
  DEEPL_API_KEY=... python3 scripts/gfood-enrich-labels.py --input docs/file.xlsx --output out.xlsx

Requires: pip install -r scripts/requirements-gfood.txt

Progress lines go to stderr every 100 rows by default (`--progress-every N`, use `0` to disable).
`--stream-output` reads the sheet in read-only mode and writes a new workbook in write-only mode
(values only; Excel formatting from the source file is not copied).
By default only rows with column **D (ACTIF)** equal to **1** or **2** are processed; use `--no-actif-filter` for all rows.
Rows with **migration_version** ≥ current target (see ``MIGRATION_VERSION_TARGET`` in script, default 1) are skipped on re-runs; enriched rows get that version written.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import unicodedata
import sqlite3
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

NEW_HEADERS = (
    "Contenance",
    "Libellé_nettoyé",
    "Marque",
    "Français",
    "Anglais",
    "Allemand",
    "Italien",
    "migration_version",
)

# Rows with migration_version >= this value are skipped (no re-enrich / no DeepL).
MIGRATION_VERSION_TARGET = 1

# Column D (0-based index 3) = ACTIF in G-Food sheet; default: only process rows with 1 or 2.
ACTIF_COL_INDEX = 3
ALLOWED_ACTIF_VALUES: frozenset[int] = frozenset({1, 2})


def actif_column_allows(value: Any) -> bool:
    """True if column D (ACTIF) is active for enrichment (1 or 2)."""
    if value is None:
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != value:  # NaN
            return False
        try:
            iv = int(round(float(value)))
        except (TypeError, ValueError, OverflowError):
            return False
        return iv in ALLOWED_ACTIF_VALUES
    s = str(value).strip()
    if s in ("1", "2"):
        return True
    try:
        iv = int(float(s.replace(",", ".")))
        return iv in ALLOWED_ACTIF_VALUES
    except ValueError:
        return False


def parse_migration_version(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != value:
            return None
        try:
            return int(round(float(value)))
        except (TypeError, ValueError, OverflowError):
            return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(float(s.replace(",", ".")))
    except ValueError:
        return None


def row_at_or_above_migration(value: Any, target: int) -> bool:
    """True if this row should be skipped (already migrated to target or newer)."""
    v = parse_migration_version(value)
    if v is None:
        return False
    return v >= target


def pad_row_list(row: list[Any], min_len: int) -> list[Any]:
    out = list(row)
    while len(out) < min_len:
        out.append(None)
    return out


def header_name_to_index(header_row: list[Any], mc: int) -> dict[str, int]:
    padded = pad_row_list(list(header_row), mc)
    m: dict[str, int] = {}
    for i, v in enumerate(padded):
        if isinstance(v, str) and v.strip():
            m[v.strip()] = i
    return m


def missing_enrichment_headers(hmap: dict[str, int]) -> list[str]:
    return [h for h in NEW_HEADERS if h not in hmap]


def open_workbook_for_row_scan(path: Path, *, data_only: bool) -> tuple[Any, Any, int, int, bool]:
    """Open workbook for scanning rows. Falls back to read_only=False when dimensions are wrong
    (common for files written with write_only=True). Returns (wb, ws, max_column, max_row, is_read_only)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=data_only)
    ws = wb.active
    mc = ws.max_column or 0
    mr = ws.max_row or 1
    # write_only saves often leave max_column=1; real G-Food sheets are wide.
    if mc < 15:
        wb.close()
        wb = openpyxl.load_workbook(path, read_only=False, data_only=data_only)
        ws = wb.active
        mc = ws.max_column or 1
        mr = ws.max_row or 1
        return wb, ws, mc, mr, False
    return wb, ws, mc, mr, True


def ensure_enrichment_columns_ws(ws: Any, new_headers: tuple[str, ...]) -> dict[str, int]:
    """Return header name -> 1-based column index; create missing headers on row 1."""
    mc = ws.max_column or 0
    name_to_col: dict[str, int] = {}
    for c in range(1, mc + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip():
            name_to_col[v.strip()] = c
    next_c = mc + 1
    for h in new_headers:
        if h not in name_to_col:
            ws.cell(row=1, column=next_c, value=h)
            name_to_col[h] = next_c
            next_c += 1
    return {h: name_to_col[h] for h in new_headers}


STOP_PRODUCT_WORDS: frozenset[str] = frozenset(
    {
        "HUILE",
        "LAIT",
        "FARINE",
        "MÉLANGE",
        "MELANGE",
        "POIVRONS",
        "TOMATES",
        "HARICOTS",
        "ACEITE",
        "OLEO",
        "ÓLEO",
        "AZEITE",
        "REIS",
        "RIZ",
        "RIJST",
        "TEIG",
        "SURIMI",
        "VINO",
        "VIN",
        "CHAMPIGNONS",
        "LÉGUMES",
        "LEGUMES",
        "CHOUX",
        "PIMENTS",
        "SAUCISSES",
        "SAUSAGE",
        "POUDRE",
        "BISCUIT",
        "PUDDING",
        "WAFFLE",
        "THÉ",
        "THE",
        "CARAMEL",
        "PAROTTA",
        "RINGE",
        "LENTILLES",
        "BULGUR",
        "FASULYE",
        "PAPRIKA",
        "SAUERKRAUT",
        "SALAT",
        "MIX",
        "GRILLÉS",
        "GRILLES",
        "MARINÉS",
        "MARINES",
        "MARINEES",
        "SECS",
        "ALIMENTAIRE",
        "ALIMENTAR",
        "ORIGINAL",
    }
)

CONNECTOR_WORDS: frozenset[str] = frozenset(
    {
        "MIT",
        "ET",
        "DE",
        "DA",
        "DI",
        "DU",
        "DES",
        "LES",
        "AU",
        "EN",
        "AUX",
        "OF",
        "AND",
        "WITH",
    }
)

LANG_TO_COL: dict[str, str] = {
    "fr": "Français",
    "en": "Anglais",
    "de": "Allemand",
    "it": "Italien",
}

# DeepL: first normalize to French from the origin segment, then translate that French to EN/DE/IT.
DEEPL_TARGETS_FROM_FR: tuple[tuple[str, str], ...] = (
    ("EN", "Anglais"),
    ("DE", "Allemand"),
    ("IT", "Italien"),
)


def normalize_ws(s: str) -> str:
    s = s.replace("\xa0", " ").replace("\u2009", " ")
    s = re.sub(r"[ \t]+", " ", s)
    return s


def strip_trailing_noise(s: str) -> str:
    s = s.rstrip()
    s = re.sub(r"\s+(PHOTO|FIXE)\s*$", "", s, flags=re.IGNORECASE)
    return s.rstrip()


def parse_decimal(num_raw: str) -> float:
    t = num_raw.strip()
    if "," in t and "." in t:
        if t.rfind(",") > t.rfind("."):
            t = t.replace(".", "").replace(",", ".")
        else:
            t = t.replace(",", "")
    elif "," in t:
        t = t.replace(",", ".")
    return float(t)


def fmt_qty(x: float) -> str:
    s = f"{x:.6f}".rstrip("0").rstrip(".")
    return s if s else "0"


def normalize_measure_token(num_raw: str, unit_raw: str, *, ge_thousand: bool) -> str:
    u = unit_raw.lower()
    qty = parse_decimal(num_raw)
    if u == "cl":
        qty *= 10.0
        u = "ml"
    elif u == "l":
        qty *= 1000.0
        u = "ml"
    elif u == "kg":
        qty *= 1000.0
        u = "g"
    elif u in ("gr", "g"):
        u = "g"

    use_ge = ge_thousand
    if u == "ml":
        if (not use_ge and qty > 1000.0) or (use_ge and qty >= 1000.0):
            return f"{fmt_qty(qty / 1000.0)} L"
        return f"{fmt_qty(qty)} ml"
    if u == "g":
        if (not use_ge and qty > 1000.0) or (use_ge and qty >= 1000.0):
            return f"{fmt_qty(qty / 1000.0)} Kg"
        return f"{fmt_qty(qty)} g"
    return f"{fmt_qty(qty)} {unit_raw}"


def normalize_contenance(raw: str, *, ge_thousand: bool) -> str:
    if not raw or not raw.strip():
        return ""
    pat = re.compile(r"(\d+[.,]?\d*)\s*(ml|cl|l|g|kg|gr)\b", re.IGNORECASE)
    parts: list[str] = []
    last = 0
    for m in pat.finditer(raw):
        if m.start() > last:
            parts.append(raw[last : m.start()])
        parts.append(normalize_measure_token(m.group(1), m.group(2), ge_thousand=ge_thousand))
        last = m.end()
    if last < len(raw):
        parts.append(raw[last:])
    return normalize_ws("".join(parts)).strip()


_TAIL_UNIT = re.compile(r"(\d+[.,]?\d*)\s*(ml|cl|l|g|kg|gr)\s*$", re.IGNORECASE)
_SEP_SUFFIX = re.compile(r"\s*(/|\s*-\s*|\s*–\s*)\s*$")


def extract_contenance(raw: str) -> tuple[str, str]:
    s = strip_trailing_noise(normalize_ws(raw))
    if not s:
        return "", ""
    i = len(s)
    start = len(s)
    while i > 0:
        chunk = s[:i]
        m = _TAIL_UNIT.search(chunk)
        if not m:
            break
        start = m.start()
        i = m.start()
        rest = s[:i].rstrip()
        if not rest:
            break
        if _SEP_SUFFIX.search(rest):
            rest2 = _SEP_SUFFIX.sub("", rest).rstrip()
            i = len(rest2)
            continue
        break
    contenance = s[start:].strip()
    label = s[:start].rstrip()
    if not contenance:
        return "", s
    return contenance, label


def clean_token_for_stop(w: str) -> str:
    return re.sub(r"^[^\w]+|[^\w]+$", "", w, flags=re.UNICODE)


def brand_lookup_key(token: str) -> str:
    core = clean_token_for_stop(token)
    if not core:
        return ""
    return unicodedata.normalize("NFC", core).casefold()


# Canonical Marque per group; variants share the same key space (NFC + casefold).
_SINGLE_TOKEN_BRAND_GROUPS: tuple[tuple[str, ...], ...] = (
    ("ÇICEK", "CICEK", "ÇIÇEK", "CIÇEK"),
    ("NESTLÉ", "NESTLE", "NESTLE\u0301"),  # decomposed é on plain NESTLE
    ("REIS",),
)


def _build_single_token_brand_map() -> dict[str, str]:
    m: dict[str, str] = {}
    for group in _SINGLE_TOKEN_BRAND_GROUPS:
        canonical = group[0]
        for variant in group:
            k = brand_lookup_key(variant)
            if k:
                m[k] = canonical
    return m


SINGLE_TOKEN_BRAND_MAP: dict[str, str] = _build_single_token_brand_map()


def guess_marque(label: str) -> tuple[str, str]:
    label = label.strip()
    if not label:
        return "", label
    tokens = label.split()
    first_key = brand_lookup_key(tokens[0])
    if first_key and first_key in SINGLE_TOKEN_BRAND_MAP:
        canonical = SINGLE_TOKEN_BRAND_MAP[first_key]
        rest = " ".join(tokens[1:]).strip()
        if not rest:
            return "", label
        return canonical, rest
    brand: list[str] = []
    for i, tok in enumerate(tokens):
        core = clean_token_for_stop(tok)
        if not core:
            break
        up = core.upper()
        if up in STOP_PRODUCT_WORDS:
            break
        if up in CONNECTOR_WORDS:
            if len(brand) <= 1:
                return "", label
            break
        if not any(c.isalpha() for c in core):
            break
        brand.append(tok)
        if len(brand) >= 4:
            break
    if not brand:
        return "", label
    rest = " ".join(tokens[len(brand) :]).strip()
    if not rest:
        return "", label
    return " ".join(brand), rest


def split_segments(clean_label: str) -> list[str]:
    t = clean_label.replace("–", "-").replace("—", "-")
    parts = re.split(r"\s+-\s+", t)
    return [normalize_ws(p).strip() for p in parts if normalize_ws(p).strip()]


def detect_lang(text: str) -> str | None:
    try:
        from langdetect import detect

        if len(text.strip()) < 3:
            return None
        return detect(text)
    except Exception:
        return None


def deepl_translate(
    texts: list[str],
    *,
    target_lang: str,
    source_lang: str | None,
    api_key: str,
    free_tier: bool,
) -> list[str]:
    base = "https://api-free.deepl.com" if free_tier else "https://api.deepl.com"
    url = f"{base}/v2/translate"
    # Header auth required since Nov 2025 (form-body auth_key deprecated).
    data: list[tuple[str, str]] = [("target_lang", target_lang)]
    if source_lang:
        data.append(("source_lang", source_lang.upper()))
    for t in texts:
        data.append(("text", t))
    body = urllib.parse.urlencode(data).encode("utf-8")
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    req.add_header("Authorization", f"DeepL-Auth-Key {api_key}")
    with urllib.request.urlopen(req, timeout=60) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    out = [item["text"] for item in payload.get("translations", [])]
    if len(out) != len(texts):
        raise RuntimeError("DeepL response length mismatch")
    return out


def cache_key(*parts: str) -> str:
    h = hashlib.sha256()
    for p in parts:
        h.update(p.encode("utf-8"))
        h.update(b"\0")
    return h.hexdigest()


def get_cache(conn: sqlite3.Connection, k: str) -> str | None:
    cur = conn.execute("SELECT v FROM t WHERE k = ?", (k,))
    row = cur.fetchone()
    return row[0] if row else None


def set_cache(conn: sqlite3.Connection, k: str, v: str) -> None:
    conn.execute("INSERT OR REPLACE INTO t(k,v) VALUES(?,?)", (k, v))
    conn.commit()


def planned_data_rows(max_row: int | None, limit: int) -> int:
    n = max(0, (max_row or 1) - 1)
    if limit and limit > 0:
        return min(limit, n)
    return n


def emit_progress(
    processed: int,
    total: int,
    t0: float,
    *,
    label: str = "[gfood]",
    actif_filter: bool = False,
    skipped_actif: int = 0,
    skipped_migrated: int = 0,
) -> None:
    elapsed = time.monotonic() - t0
    if actif_filter:
        mig = f" | {skipped_migrated} skipped (already migrated)" if skipped_migrated else ""
        print(
            f"{label} {processed} rows kept | {skipped_actif} skipped (column D ACTIF not 1/2){mig} "
            f"| {elapsed:.0f}s elapsed",
            file=sys.stderr,
            flush=True,
        )
        return
    if total > 0:
        pct = 100.0 * processed / total
        if processed > 0 and elapsed > 0:
            rate = processed / elapsed
            eta = (total - processed) / rate if rate > 0 else 0.0
            extra = f" ~{eta:.0f}s left"
        else:
            extra = ""
        mig = f" | {skipped_migrated} skip-migrated" if skipped_migrated else ""
        print(
            f"{label} {processed}/{total} rows ({pct:.1f}%) | {elapsed:.0f}s elapsed{extra}{mig}",
            file=sys.stderr,
            flush=True,
        )
    else:
        mig = f" | {skipped_migrated} skip-migrated" if skipped_migrated else ""
        print(
            f"{label} {processed} rows | {elapsed:.0f}s elapsed{mig}",
            file=sys.stderr,
            flush=True,
        )


def ensure_cache(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(path))
    conn.execute("CREATE TABLE IF NOT EXISTS t (k TEXT PRIMARY KEY, v TEXT NOT NULL)")
    conn.commit()
    return conn


def map_segment_langs(segments: list[str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for seg in segments:
        if len(seg.strip()) < 28:
            continue
        lang = detect_lang(seg)
        if not lang:
            continue
        col = LANG_TO_COL.get(lang)
        if not col:
            continue
        prev = out.get(col)
        if prev is None or len(seg) > len(prev):
            out[col] = seg
    return out


def deepl_source_code(detected: str | None) -> str | None:
    if not detected:
        return None
    m = {"fr": "FR", "en": "EN", "de": "DE", "it": "IT"}
    return m.get(detected)


def process_row(
    raw_libelle: str,
    *,
    ge_thousand: bool,
    translate: bool,
    api_key: str | None,
    cache: sqlite3.Connection | None,
    free_tier: bool,
) -> dict[str, Any]:
    contenance_raw, label_wo = extract_contenance(raw_libelle or "")
    contenance = normalize_contenance(contenance_raw, ge_thousand=ge_thousand)
    marque, titre = guess_marque(label_wo)
    display_label = titre if marque else label_wo
    segments = split_segments(display_label)
    origin = segments[0] if segments else display_label

    langs_from_segments = map_segment_langs(segments) if len(segments) > 1 else {}
    detected = detect_lang(origin)
    src = deepl_source_code(detected)

    cols: dict[str, Any] = {h: "" for h in NEW_HEADERS}
    cols["Contenance"] = contenance
    cols["Libellé_nettoyé"] = display_label
    cols["Marque"] = marque

    if not translate or not api_key:
        for k, v in langs_from_segments.items():
            cols[k] = v
        return cols

    # DeepL path: ignore segment-based language guesses; French is the hub.
    text_to_translate = origin.strip()
    if not text_to_translate:
        return cols

    def one_call(
        text: str,
        *,
        target_lang: str,
        source_lang: str | None,
        cache_tag: str,
    ) -> str:
        ck = cache_key("deepl", cache_tag, source_lang or "auto", target_lang, text)
        if cache is not None:
            hit = get_cache(cache, ck)
            if hit is not None:
                return hit
        try:
            out = deepl_translate(
                [text],
                target_lang=target_lang,
                source_lang=source_lang,
                api_key=api_key,
                free_tier=free_tier,
            )[0]
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"DeepL HTTP {e.code}: {e.read()[:500]!r}") from e
        if cache is not None:
            set_cache(cache, ck, out)
        time.sleep(0.05)
        return out

    # 1) Reference French (translate origin → FR unless already French)
    if detected == "fr":
        french_text = text_to_translate
    else:
        french_text = one_call(
            text_to_translate,
            target_lang="FR",
            source_lang=src,
            cache_tag="hub2fr",
        )
    cols["Français"] = french_text

    # 2) EN / DE / IT always from French
    for deepl_tgt, header in DEEPL_TARGETS_FROM_FR:
        cols[header] = one_call(
            french_text,
            target_lang=deepl_tgt,
            source_lang="FR",
            cache_tag="fr2",
        )

    return cols


def finalize_enriched_row(cols: dict[str, Any]) -> None:
    cols["migration_version"] = MIGRATION_VERSION_TARGET


def run_self_test() -> None:
    assert normalize_contenance("1700 ml", ge_thousand=False) == "1.7 L"
    assert normalize_contenance("1000 ml", ge_thousand=False) == "1000 ml"
    assert normalize_contenance("1000 ml", ge_thousand=True) == "1 L"
    assert normalize_contenance("50cl", ge_thousand=False) == "500 ml"
    c_raw, lab = extract_contenance("ROMAGE BLANC CRÉMEUX DES BALKANS 800G / 1.5KG")
    assert "800" in c_raw and "1.5" in c_raw
    assert "ROMAGE" in lab
    n = normalize_contenance(c_raw, ge_thousand=False)
    assert "g" in n and "Kg" in n
    m, r = guess_marque("NESTLÉ NIDO LAIT EN POUDRE")
    assert m == "NESTLÉ"
    assert r.startswith("NIDO")
    m2, r2 = guess_marque("NESTLE NIDO LAIT EN POUDRE")
    assert m2 == "NESTLÉ"
    m3, r3 = guess_marque("ÇICEK KARISIK SEBZE TURSUSU - MÉLANGE 3KG")
    assert m3 == "ÇICEK"
    assert "KARISIK" in r3
    m4, r4 = guess_marque("REIS LENTILLES ROUGES 2.5KG")
    assert m4 == "REIS"
    assert "LENTILLES" in r4
    assert actif_column_allows(1) and actif_column_allows(2)
    assert actif_column_allows(1.0) and actif_column_allows("2")
    assert not actif_column_allows(0) and not actif_column_allows(3) and not actif_column_allows(None)
    assert parse_migration_version(1) == 1 and parse_migration_version("1") == 1
    assert parse_migration_version(None) is None
    assert row_at_or_above_migration(1, 1) and row_at_or_above_migration(2, 1)
    assert not row_at_or_above_migration(0, 1) and not row_at_or_above_migration(None, 1)
    print("self-test OK", file=sys.stderr)


def main() -> int:
    ap = argparse.ArgumentParser(description="Enrich G-Food Libellé column in Excel.")
    ap.add_argument("--input", "-i", type=Path, default=None, help="Source .xlsx path")
    ap.add_argument(
        "--output",
        "-o",
        type=Path,
        default=None,
        help="Output .xlsx (required unless --dry-run)",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Do not write xlsx; print JSON summary and samples (no DeepL unless --dry-run-translate)",
    )
    ap.add_argument(
        "--dry-run-translate",
        action="store_true",
        help="With --dry-run, call DeepL when DEEPL_API_KEY is set (still no xlsx; uses translation cache)",
    )
    ap.add_argument("--limit", type=int, default=0, help="Max data rows to process (0 = all)")
    ap.add_argument(
        "--ge-thousand",
        action="store_true",
        help="Use >= 1000 ml/g for L/Kg conversion (default: strict > 1000)",
    )
    ap.add_argument(
        "--deepl-free",
        action="store_true",
        help="Use api-free.deepl.com (default: pro api.deepl.com)",
    )
    ap.add_argument(
        "--progress-every",
        type=int,
        default=100,
        metavar="N",
        help="Print progress to stderr every N data rows (0 = off). Default 100.",
    )
    ap.add_argument(
        "--stream-output",
        action="store_true",
        help="Read input in read-only mode and write with write_only workbook (lower memory; "
        "cell formatting is not preserved). Requires --output; not for --dry-run.",
    )
    ap.add_argument(
        "--no-actif-filter",
        action="store_true",
        help="Process every data row. Default: skip rows where column D (ACTIF) is not 1 or 2.",
    )
    ap.add_argument("--self-test", action="store_true", help="Run inline assertions and exit")
    args = ap.parse_args()

    if args.self_test:
        run_self_test()
        return 0

    if args.input is None:
        ap.error("--input/-i is required unless --self-test")

    if args.dry_run_translate and not args.dry_run:
        ap.error("--dry-run-translate requires --dry-run")

    if not args.dry_run and args.output is None:
        ap.error("--output is required unless --dry-run")

    if args.stream_output and args.dry_run:
        ap.error("--stream-output cannot be used with --dry-run")

    try:
        import openpyxl
    except ImportError:
        print("Install deps: pip install -r scripts/requirements-gfood.txt", file=sys.stderr)
        return 1

    api_key = os.environ.get("DEEPL_API_KEY")
    if not api_key and (not args.dry_run or args.dry_run_translate):
        if args.dry_run_translate:
            print(
                "Warning: DEEPL_API_KEY not set; --dry-run-translate will not call DeepL.",
                file=sys.stderr,
            )
        elif not args.dry_run:
            print(
                "Warning: DEEPL_API_KEY not set; translation columns will stay empty.",
                file=sys.stderr,
            )

    cache_path = Path(__file__).resolve().parent / ".gfood-translate-cache.sqlite"
    use_cache = bool(api_key) and (not args.dry_run or args.dry_run_translate)
    cache_conn = ensure_cache(cache_path) if use_cache else None
    translate = bool(api_key) and (not args.dry_run or args.dry_run_translate)
    actif_filter = not args.no_actif_filter

    stats: dict[str, Any] = {
        "rows": 0,
        "with_contenance": 0,
        "with_marque": 0,
        "skipped_actif": 0,
        "skipped_migrated": 0,
        "migration_version_target": MIGRATION_VERSION_TARGET,
    }
    samples: list[dict[str, Any]] = []

    def collect_stats(row_num: int, libelle: str, cols: dict[str, Any]) -> None:
        stats["rows"] += 1
        if cols["Contenance"]:
            stats["with_contenance"] += 1
        if cols["Marque"]:
            stats["with_marque"] += 1
        if len(samples) < 25:
            samples.append({"row": row_num, "raw": libelle[:120], **cols})

    def run_process_row(libelle: str) -> dict[str, Any]:
        return process_row(
            libelle,
            ge_thousand=args.ge_thousand,
            translate=translate,
            api_key=api_key,
            cache=cache_conn,
            free_tier=args.deepl_free,
        )

    if args.dry_run:
        wb, ws, mc, mr, _ = open_workbook_for_row_scan(args.input, data_only=True)
        total = planned_data_rows(mr, args.limit)
        t0 = time.monotonic()
        if args.progress_every:
            filt = "ACTIF column D ∈ {{1,2}} (use --no-actif-filter for all rows)" if actif_filter else "no ACTIF filter"
            print(
                f"[gfood] dry-run up to {total} matching rows ({filt}); progress every {args.progress_every}",
                file=sys.stderr,
                flush=True,
            )
        rows_iter = ws.iter_rows(
            min_row=1, max_row=mr, min_col=1, max_col=mc, values_only=True
        )
        try:
            header_row = list(next(rows_iter))
        except StopIteration:
            print("Empty sheet", file=sys.stderr)
            wb.close()
            return 1
        try:
            libelle_idx = header_row.index("Libellé")
        except ValueError:
            print('Header row must contain column "Libellé"', file=sys.stderr)
            wb.close()
            return 1
        hmap = header_name_to_index(header_row, mc)
        migration_idx = hmap.get("migration_version")
        processed = 0
        row_num = 1
        for row in rows_iter:
            row_num += 1
            if args.limit and processed >= args.limit:
                break
            row_list = list(row)
            if len(row_list) < mc:
                row_list.extend([None] * (mc - len(row_list)))
            if actif_filter:
                av = row_list[ACTIF_COL_INDEX] if len(row_list) > ACTIF_COL_INDEX else None
                if not actif_column_allows(av):
                    stats["skipped_actif"] += 1
                    continue
            if migration_idx is not None:
                mv = row_list[migration_idx] if migration_idx < len(row_list) else None
                if row_at_or_above_migration(mv, MIGRATION_VERSION_TARGET):
                    stats["skipped_migrated"] += 1
                    continue
            libelle_cell = row_list[libelle_idx] if libelle_idx < len(row_list) else None
            libelle = "" if libelle_cell is None else str(libelle_cell)
            cols = run_process_row(libelle)
            finalize_enriched_row(cols)
            collect_stats(row_num, libelle, cols)
            processed += 1
            if args.progress_every and processed % args.progress_every == 0:
                emit_progress(
                    processed,
                    total,
                    t0,
                    actif_filter=actif_filter,
                    skipped_actif=stats["skipped_actif"],
                    skipped_migrated=stats["skipped_migrated"],
                )
        wb.close()
        if args.progress_every and processed > 0:
            if processed % args.progress_every != 0:
                emit_progress(
                    processed,
                    total,
                    t0,
                    actif_filter=actif_filter,
                    skipped_actif=stats["skipped_actif"],
                    skipped_migrated=stats["skipped_migrated"],
                )
            print(
                f"[gfood] finished {processed} rows in {time.monotonic() - t0:.1f}s",
                file=sys.stderr,
                flush=True,
            )
    elif args.stream_output:
        assert args.output is not None
        wb_in, ws_in, mc, mr, _ = open_workbook_for_row_scan(args.input, data_only=True)
        total = planned_data_rows(mr, args.limit)
        t0 = time.monotonic()
        if args.progress_every:
            filt = "ACTIF D∈{{1,2}} only" if actif_filter else "all rows"
            print(
                f"[gfood] stream-output → {args.output} (up to {total} kept rows, {filt}; "
                f"every {args.progress_every}; formatting not copied)",
                file=sys.stderr,
                flush=True,
            )
        rows_it = ws_in.iter_rows(
            min_row=1, max_row=mr, min_col=1, max_col=mc, values_only=True
        )
        try:
            header_row = list(next(rows_it))
        except StopIteration:
            print("Empty sheet", file=sys.stderr)
            wb_in.close()
            return 1
        try:
            libelle_idx = header_row.index("Libellé")
        except ValueError:
            print('Header row must contain column "Libellé"', file=sys.stderr)
            wb_in.close()
            return 1
        hmap = header_name_to_index(header_row, mc)
        missing_tail = missing_enrichment_headers(hmap)
        migration_idx = hmap.get("migration_version")
        hdr_out = pad_row_list(list(header_row), mc) + missing_tail
        wb_out = openpyxl.Workbook(write_only=True)
        ws_out = wb_out.create_sheet(title=ws_in.title or "Sheet")
        ws_out.append(hdr_out)
        processed = 0
        row_num = 1
        try:
            for row in rows_it:
                row_num += 1
                if args.limit and processed >= args.limit:
                    break
                row_list = list(row)
                if len(row_list) < mc:
                    row_list.extend([None] * (mc - len(row_list)))
                if actif_filter:
                    av = row_list[ACTIF_COL_INDEX] if len(row_list) > ACTIF_COL_INDEX else None
                    if not actif_column_allows(av):
                        stats["skipped_actif"] += 1
                        continue
                if migration_idx is not None:
                    mv = row_list[migration_idx] if migration_idx < len(row_list) else None
                    if row_at_or_above_migration(mv, MIGRATION_VERSION_TARGET):
                        stats["skipped_migrated"] += 1
                        out_pass = pad_row_list(list(row_list), mc)
                        while len(out_pass) < len(hdr_out):
                            out_pass.append(None)
                        ws_out.append(out_pass[: len(hdr_out)])
                        continue
                libelle_cell = row_list[libelle_idx] if libelle_idx < len(row_list) else None
                libelle = "" if libelle_cell is None else str(libelle_cell)
                cols = run_process_row(libelle)
                finalize_enriched_row(cols)
                collect_stats(row_num, libelle, cols)
                out_row = pad_row_list(list(row_list), mc)
                for h in NEW_HEADERS:
                    if h in hmap:
                        idx = hmap[h]
                        while len(out_row) <= idx:
                            out_row.append(None)
                        out_row[idx] = cols[h]
                for h in missing_tail:
                    out_row.append(cols[h])
                ws_out.append(out_row)
                processed += 1
                if args.progress_every and processed % args.progress_every == 0:
                    emit_progress(
                        processed,
                        total,
                        t0,
                        actif_filter=actif_filter,
                        skipped_actif=stats["skipped_actif"],
                        skipped_migrated=stats["skipped_migrated"],
                    )
        finally:
            wb_in.close()
        if args.progress_every and processed > 0:
            if processed % args.progress_every != 0:
                emit_progress(
                    processed,
                    total,
                    t0,
                    actif_filter=actif_filter,
                    skipped_actif=stats["skipped_actif"],
                    skipped_migrated=stats["skipped_migrated"],
                )
            print(
                f"[gfood] finished {processed} rows in {time.monotonic() - t0:.1f}s → save…",
                file=sys.stderr,
                flush=True,
            )
        wb_out.save(args.output)
    else:
        wb = openpyxl.load_workbook(args.input, read_only=False, data_only=True)
        ws = wb.active
        header_row = [ws.cell(row=1, column=c).value for c in range(1, (ws.max_column or 0) + 1)]
        while header_row and header_row[-1] is None:
            header_row.pop()
        if not header_row:
            print("Empty sheet", file=sys.stderr)
            return 1
        try:
            libelle_idx = header_row.index("Libellé")
        except ValueError:
            print('Header row must contain column "Libellé"', file=sys.stderr)
            return 1
        col_map = ensure_enrichment_columns_ws(ws, NEW_HEADERS)
        migration_col = col_map["migration_version"]
        processed = 0
        max_row = ws.max_row or 1
        total = planned_data_rows(max_row, args.limit)
        t0 = time.monotonic()
        if args.progress_every:
            filt = "ACTIF D∈{{1,2}} only" if actif_filter else "all rows"
            print(
                f"[gfood] in-place write (up to {total} kept rows, {filt}; "
                f"migration_target={MIGRATION_VERSION_TARGET}; every {args.progress_every})",
                file=sys.stderr,
                flush=True,
            )
        for row_num in range(2, max_row + 1):
            if args.limit and processed >= args.limit:
                break
            if actif_filter:
                actif_cell = ws.cell(row=row_num, column=4).value
                if not actif_column_allows(actif_cell):
                    stats["skipped_actif"] += 1
                    continue
            mv_cell = ws.cell(row=row_num, column=migration_col).value
            if row_at_or_above_migration(mv_cell, MIGRATION_VERSION_TARGET):
                stats["skipped_migrated"] += 1
                continue
            libelle_cell = ws.cell(row=row_num, column=libelle_idx + 1).value
            libelle = "" if libelle_cell is None else str(libelle_cell)
            cols = run_process_row(libelle)
            finalize_enriched_row(cols)
            collect_stats(row_num, libelle, cols)
            for name in NEW_HEADERS:
                ws.cell(row=row_num, column=col_map[name], value=cols[name])
            processed += 1
            if args.progress_every and processed % args.progress_every == 0:
                emit_progress(
                    processed,
                    total,
                    t0,
                    actif_filter=actif_filter,
                    skipped_actif=stats["skipped_actif"],
                    skipped_migrated=stats["skipped_migrated"],
                )
        assert args.output is not None
        if args.progress_every and processed > 0:
            if processed % args.progress_every != 0:
                emit_progress(
                    processed,
                    total,
                    t0,
                    actif_filter=actif_filter,
                    skipped_actif=stats["skipped_actif"],
                    skipped_migrated=stats["skipped_migrated"],
                )
            print(
                f"[gfood] finished {processed} rows in {time.monotonic() - t0:.1f}s → save…",
                file=sys.stderr,
                flush=True,
            )
        wb.save(args.output)
        wb.close()

    print(json.dumps({"stats": stats, "samples": samples}, ensure_ascii=False, indent=2))
    if args.dry_run:
        if args.dry_run_translate and api_key:
            print(
                "\n(dry-run: no xlsx written; DeepL filled JSON samples / cache)",
                file=sys.stderr,
            )
        elif args.dry_run_translate:
            print(
                "\n(dry-run: no xlsx written; translations skipped — no DEEPL_API_KEY)",
                file=sys.stderr,
            )
        else:
            print("\n(dry-run: no file written, no DeepL calls)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
